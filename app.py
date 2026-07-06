import cv2
import platform
import threading
import tkinter as tk
from tkinter import messagebox, filedialog
from PIL import Image, ImageTk
import os
import subprocess
import io
import openpyxl
from google.cloud import vision

if "GOOGLE_APPLICATION_CREDENTIALS" not in os.environ:
    print("Error: GOOGLE_APPLICATION_CREDENTIALS environment variable not set.")

FIELD_COORDINATES = {
    "name": (89, 11, 275, 46),
    "usn": (374, 22, 553, 54),
    "course_code": (415, 57, 550, 96),
    "mse1_unit1": (211, 231, 268, 302),
    "mse1_unit2": (277, 236, 335, 303),
}

def load_image(image_path):
    return Image.open(image_path)

def extract_text_from_region(image, coords, client):
    cropped_image = image.crop(coords)
    img_byte_arr = io.BytesIO()
    cropped_image.save(img_byte_arr, format='PNG')
    content = img_byte_arr.getvalue()

    image_for_vision = vision.Image(content=content)
    response = client.text_detection(image=image_for_vision)

    if response.error.message:
        raise Exception(f"Vision API Error: {response.error.message}")

    texts = response.text_annotations
    return texts[0].description.strip() if texts else ""

def extract_fields(image_path):
    client = vision.ImageAnnotatorClient()
    image = load_image(image_path)
    extracted_data = {}
    for field, coords in FIELD_COORDINATES.items():
        text = extract_text_from_region(image, coords, client)
        if field == "name":
            text = text.replace("\n", " ")
        else:
            text = text.replace(" ", "").replace("\n", "")
        extracted_data[field] = text
    return extracted_data

class DroidCamGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("DroidCam Live Feed")
        self.root.geometry("780x600")

        main_frame = tk.Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(main_frame)
        scrollbar = tk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        tk.Label(self.scrollable_frame, text="DroidCam URL:").pack()
        self.url_entry = tk.Entry(self.scrollable_frame, width=50)
        self.url_entry.pack()
        self.url_entry.insert(0, "http://192.168.29.148:4747/video")

        self.start_button = tk.Button(self.scrollable_frame, text="Start", command=self.start_video)
        self.start_button.pack(pady=5)

        folder_frame = tk.Frame(self.scrollable_frame)
        folder_frame.pack(pady=5)
        tk.Label(folder_frame, text="Enter Folder Name:").pack(side=tk.LEFT, padx=5)
        self.folder_entry = tk.Entry(folder_frame, width=25)
        self.folder_entry.pack(side=tk.LEFT, padx=5)
        self.browse_button = tk.Button(folder_frame, text="Browse Folder", command=self.open_folder)
        self.browse_button.pack(side=tk.LEFT, padx=5)

        self.video_label = tk.Label(self.scrollable_frame)
        self.video_label.pack(pady=5)

        image_frame = tk.Frame(self.scrollable_frame)
        image_frame.pack(pady=10)
        tk.Label(image_frame, text="Image Name:").pack(side=tk.LEFT, padx=5)
        self.image_name_entry = tk.Entry(image_frame, width=25)
        self.image_name_entry.pack(side=tk.LEFT, padx=5)
        self.save_button = tk.Button(image_frame, text="Save", command=self.save_frame)
        self.save_button.pack(side=tk.LEFT, padx=5)

        browse_frame = tk.Frame(self.scrollable_frame)
        browse_frame.pack(pady=10)
        self.browse_image_button = tk.Button(browse_frame, text="Browse Image", command=self.browse_image_file)
        self.browse_image_button.pack(side=tk.LEFT, padx=5)
        self.selected_image_label = tk.Label(browse_frame, text="No file selected")
        self.selected_image_label.pack(side=tk.LEFT, padx=5)
        self.extract_button = tk.Button(browse_frame, text="Extract", command=self.extract_selected_image)
        self.extract_button.pack(side=tk.LEFT, padx=5)

        self.text_box = tk.Text(self.scrollable_frame, width=90, height=10)
        self.text_box.pack(pady=10)

        excel_frame = tk.Frame(self.scrollable_frame)
        excel_frame.pack(pady=10)
        tk.Label(excel_frame, text="Excel File Name:").pack(side=tk.LEFT, padx=5)
        self.excel_name_entry = tk.Entry(excel_frame, width=25)
        self.excel_name_entry.pack(side=tk.LEFT, padx=5)
        self.export_button = tk.Button(excel_frame, text="Export to Excel", command=self.export_to_excel)
        self.export_button.pack(side=tk.LEFT, padx=5)

        self.open_excel_button = tk.Button(self.scrollable_frame, text="Open Excel File", command=self.open_excel_file)
        self.open_excel_button.pack(pady=5)

        self.cap = None
        self.running = False
        self.frame = None
        self.last_saved_path = ""
        self.selected_image_path = ""
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def get_folder_path(self):
        folder_name = self.folder_entry.get().strip()
        if not folder_name:
            folder_name = "DroidCamFrames"
        folder_path = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", folder_name)
        os.makedirs(folder_path, exist_ok=True)
        return folder_path

    def start_video(self):
        url = self.url_entry.get().strip()
        if not url:
            messagebox.showerror("Error", "Please enter a DroidCam URL")
            return

        self.cap = cv2.VideoCapture(url)
        if not self.cap.isOpened():
            messagebox.showerror("Error", "Unable to open video stream")
            return

        self.running = True
        # REMOVED threading. Started Tkinter loop instead.
        self.update_frame()

    def update_frame(self):
        if self.running and self.cap.isOpened():
            ret, frame = self.cap.read()
            if ret:
                self.frame = frame
                rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                image = Image.fromarray(rgb_frame).resize((640, 480))
                photo = ImageTk.PhotoImage(image=image)

                self.video_label.configure(image=photo)
                self.video_label.image = photo
            
            # Schedules the next frame update safely on the main GUI thread in 15 milliseconds
            self.root.after(15, self.update_frame)
        else:
            if self.cap:
                self.cap.release()

    def save_frame(self):
        if self.frame is None:
            messagebox.showwarning("No Frame", "No frame available to save.")
            return

        image_name = self.image_name_entry.get().strip()
        if not image_name:
            messagebox.showwarning("Missing Name", "Please enter an image name.")
            return

        folder_path = self.get_folder_path()
        file_path = os.path.join(folder_path, image_name + ".jpg")
        try:
            cv2.imwrite(file_path, self.frame)
            self.last_saved_path = file_path
            messagebox.showinfo("Saved", f"Saved at: {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save image:\n{e}")

    def browse_image_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.png *.jpeg")])
        if file_path:
            self.selected_image_path = file_path
            self.selected_image_label.config(text=os.path.basename(file_path))

    def extract_selected_image(self):
        if not self.selected_image_path and not self.last_saved_path:
            messagebox.showwarning("No Image", "Please select or save an image first.")
            return

        try:
            path_to_use = self.selected_image_path if self.selected_image_path else self.last_saved_path
            extracted = extract_fields(path_to_use)
            result = "\n".join(f"{k.capitalize()}: {v}" for k, v in extracted.items())
            self.text_box.delete("1.0", tk.END)
            self.text_box.insert(tk.END, result)
        except Exception as e:
            messagebox.showerror("Error", f"Extraction failed:\n{e}")

    def export_to_excel(self):
        text = self.text_box.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("No Data", "No extracted data to export.")
            return

        excel_name = self.excel_name_entry.get().strip()
        if not excel_name:
            messagebox.showwarning("Missing Name", "Please enter an Excel file name.")
            return

        folder_path = self.get_folder_path()
        excel_path = os.path.join(folder_path, excel_name + ".xlsx")

        extracted_data = {}
        for line in text.splitlines():
            key_val = line.split(":", 1)
            if len(key_val) == 2:
                key, val = key_val
                extracted_data[key.strip().lower()] = val.strip()

        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            next_row = ws.max_row + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["name", "usn", "course_code", "mse1_unit1", "mse1_unit2"]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header.capitalize())
            next_row = 2

        headers = ["name", "usn", "course_code", "mse1_unit1", "mse1_unit2"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=next_row, column=col_idx, value=extracted_data.get(header, ""))

        wb.save(excel_path)
        messagebox.showinfo("Exported", f"Data exported to {excel_path}")
        self.last_excel_path = excel_path

    def open_excel_file(self):
        try:
            if hasattr(self, 'last_excel_path') and os.path.exists(self.last_excel_path):
                # Cross-platform way to open files
                if platform.system() == 'Windows':
                    os.startfile(self.last_excel_path)
                elif platform.system() == 'Darwin': # macOS
                    subprocess.call(('open', self.last_excel_path))
                else: # Linux
                    subprocess.call(('xdg-open', self.last_excel_path))
            else:
                messagebox.showwarning("Not Found", "No Excel file found. Export first.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel file:\n{e}")

    def open_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            folder_name = os.path.basename(folder_selected)
            self.folder_entry.delete(0, tk.END)
            self.folder_entry.insert(0, folder_name)
            try:
                # Cross-platform way to open folders
                if platform.system() == 'Windows':
                    os.startfile(folder_selected)
                elif platform.system() == 'Darwin':
                    subprocess.call(('open', folder_selected))
                else:
                    subprocess.call(('xdg-open', folder_selected))
            except Exception as e:
                messagebox.showerror("Error", f"Could not open folder:\n{e}")

    def on_closing(self):
        self.running = False
        if self.cap:
            self.cap.release()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = DroidCamGUI(root)
    root.mainloop()

